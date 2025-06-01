import string
import random
import csv
import io
from flask import render_template, request, redirect, url_for, flash, session, make_response, Blueprint
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from models import User, ActivityLog, db
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

routes_bp = Blueprint('routes_bp', __name__)

@routes_bp.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('routes_bp.generator'))
    return redirect(url_for('routes_bp.login'))

@routes_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash('Logged in successfully!', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('routes_bp.generator'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('login.html')



@routes_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('routes_bp.login'))

@routes_bp.route('/generator', methods=['GET', 'POST'])
@login_required
def generator():
    if request.method == 'POST':
        # Get form data
        base_name = request.form['base_name']
        base_ip = request.form['base_ip']
        comment = request.form['comment']
        start_number = int(request.form['start_number'])
        end_number = int(request.form['end_number'])
        password_length = int(request.form['password_length'])
        
        # Character types
        char_types = []
        if 'uppercase' in request.form:
            char_types.append('uppercase')
        if 'lowercase' in request.form:
            char_types.append('lowercase')
        if 'numbers' in request.form:
            char_types.append('numbers')
        if 'special' in request.form:
            char_types.append('special')
        
        # Validation
        if not char_types:
            flash('Please select at least one character type.', 'error')
            return render_template('generator.html')
        
        if start_number > end_number:
            flash('Start number cannot be greater than end number.', 'error')
            return render_template('generator.html')
        
        # Validate IP range (1-254 for last octet)
        if start_number < 1 or end_number > 254:
            flash('IP range must be between 1 and 254.', 'error')
            return render_template('generator.html')
        
        # Generate character set
        charset = ''
        if 'uppercase' in char_types:
            charset += string.ascii_uppercase
        if 'lowercase' in char_types:
            charset += string.ascii_lowercase
        if 'numbers' in char_types:
            charset += string.digits
        if 'special' in char_types:
            charset += '!@#$%^&*'
        
        # Generate commands (optimized for speed)
        commands = []
        generated_users = []  # Store user data for Excel export
        users_count = end_number - start_number + 1
        
        # Pre-calculate IP base parts for efficiency
        ip_parts = base_ip.split('.')
        is_partial_ip = len(ip_parts) == 3
        base_ip_int = int(ip_parts[3]) if len(ip_parts) == 4 else 0
        
        # Generate all users in one loop (optimized)
        for i in range(start_number, end_number + 1):
            # Generate random password (optimized)
            password = ''.join(random.choices(charset, k=password_length))
            
            # Calculate IP address efficiently
            if is_partial_ip:
                user_ip = f"{base_ip}.{i}"
            elif len(ip_parts) == 4:
                user_ip = f"{ip_parts[0]}.{ip_parts[1]}.{ip_parts[2]}.{base_ip_int + i - start_number}"
            else:
                user_ip = base_ip
            
            username = f"{base_name}{i}"
            
            # Store user data for Excel export
            generated_users.append({
                'name': username,
                'password': password,
                'ip': user_ip,
                'comment': comment
            })
            
            # Generate command in the new format
            command = f' add comment={comment} address={user_ip} name={username} password={password}'
            commands.append(command)
        
        # Log activity
        activity = ActivityLog(
            user_id=current_user.id,
            base_name=base_name,
            base_ip=base_ip,
            comment=comment,
            start_number=start_number,
            end_number=end_number,
            password_length=password_length,
            character_types=','.join(char_types),
            users_generated=users_count
        )
        
        db.session.add(activity)
        db.session.commit()
        
        # Store generated users in session for Excel export
        session['generated_users'] = generated_users
        session['export_metadata'] = {
            'base_name': base_name,
            'comment': comment,
            'users_count': users_count,
            'generated_by': current_user.username,
            'export_date': datetime.now().isoformat()
        }
        
        # Format the commands with the header
        commands_text = '/ip hotspot user\n' + '\n'.join(commands)
        
        return render_template('generator.html', commands=commands, commands_text=commands_text)
    
    return render_template('generator.html')

@routes_bp.route('/admin')
@login_required
def admin():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('routes_bp.generator'))
    
    users = User.query.all()
    recent_activities = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).limit(10).all()
    
    return render_template('admin.html', users=users, recent_activities=recent_activities)

@routes_bp.route('/admin/create_user', methods=['POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('routes_bp.generator'))
    
    username = request.form['username']
    email = request.form['email']
    password = request.form['password']
    confirm_password = request.form['confirm_password']
    is_admin = 'is_admin' in request.form
    
    # Validation
    if password != confirm_password:
        flash('Passwords do not match.', 'error')
        return redirect(url_for('routes_bp.admin'))
    
    if User.query.filter_by(username=username).first():
        flash('Username already exists.', 'error')
        return redirect(url_for('routes_bp.admin'))
    
    if User.query.filter_by(email=email).first():
        flash('Email already registered.', 'error')
        return redirect(url_for('routes_bp.admin'))
    
    # Create new user
    user = User(
        username=username,
        email=email,
        password_hash=generate_password_hash(password),
        is_admin=is_admin
    )
    
    db.session.add(user)
    db.session.commit()
    
    flash(f'User "{username}" created successfully!', 'success')
    return redirect(url_for('routes_bp.admin'))

@routes_bp.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('routes_bp.generator'))
    
    user = User.query.get_or_404(user_id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('routes_bp.admin'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    flash(f'User "{username}" deleted successfully!', 'success')
    return redirect(url_for('routes_bp.admin'))

@routes_bp.route('/activity')
@login_required
def activity():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    if current_user.is_admin:
        # Admin can see all activities
        activities = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        # Regular users only see their own activities
        activities = ActivityLog.query.filter_by(user_id=current_user.id).order_by(ActivityLog.timestamp.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
        
    return render_template('activity.html', activities=activities)

@routes_bp.route('/export_activity')
@login_required
def export_activity():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('routes_bp.generator'))
    
    activities = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).all()
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Log"
    
    # Add header row with styling
    header = ['Timestamp', 'User', 'Base Name', 'Base IP', 'Comment', 'Start #', 'End #', 'Password Length', 'Char Types', 'Users Generated']
    ws.append(header)
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light grey fill
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col)].width = 15 # Set a default width
    
    # Add data rows
    for activity in activities:
        ws.append([
            activity.timestamp,
            activity.user.username,
            activity.base_name,
            activity.base_ip,
            activity.comment,
            activity.start_number,
            activity.end_number,
            activity.password_length,
            activity.character_types,
            activity.users_generated
        ])
    
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column index
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2 # Add some padding
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    
    # Create a response for the Excel file
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=activity_log_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@routes_bp.route('/export_users_excel')
@login_required
def export_users_excel():
    # Check if there are generated users in session
    generated_users = session.get('generated_users')
    export_metadata = session.get('export_metadata')
    
    if not generated_users:
        flash('No users generated in the current session to export.', 'warning')
        return redirect(url_for('routes_bp.generator'))
        
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Generated Users"
    
    # Add metadata (optional)
    if export_metadata:
        ws.append([f'Generated by: {export_metadata.get('generated_by', 'N/A')}'])
        ws.append([f'Generation Date: {export_metadata.get('export_date', 'N/A')}'])
        ws.append([f'Base Name: {export_metadata.get('base_name', 'N/A')}'])
        ws.append([f'Comment: {export_metadata.get('comment', 'N/A')}'])
        ws.append([]) # Add an empty row for spacing

    
    # Add header row with styling
    header = ['Name', 'Password', 'IP Address', 'Comment']
    ws.append(header)
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Light grey fill
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col)].width = 15 # Set a default width
        
    # Add data rows
    for user_data in generated_users:
        ws.append([
            user_data['name'],
            user_data['password'],
            user_data['ip'],
            user_data['comment']
        ])
        
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column # Get the column index
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2 # Add some padding
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    
    # Create a response for the Excel file
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = make_response(excel_file.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=generated_users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response
